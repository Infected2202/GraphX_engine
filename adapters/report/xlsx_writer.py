"""Excel report writer."""
from __future__ import annotations

from pathlib import Path
from typing import Sequence

try:  # pragma: no cover - exercised in integration
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
except ImportError as exc:  # pragma: no cover - depends on environment
    Workbook = None
    Alignment = Font = None
    _IMPORT_ERROR = exc
else:  # pragma: no cover - imported when dependency present
    _IMPORT_ERROR = None

from domain.models import Employee
from domain.schedule import Schedule


class XLSXExportUnavailable(RuntimeError):
    """Raised when Excel export cannot run due to missing dependencies."""


HEADER_FONT = Font(bold=True) if Font else None
CENTER = Alignment(horizontal="center", vertical="center") if Alignment else None


def write_grid(path: str | Path, schedule: Schedule, employees: Sequence[Employee], *, title: str | None = None) -> Path:
    if Workbook is None:
        raise XLSXExportUnavailable(
            "openpyxl is required for Excel export. Install 'openpyxl' to enable XLSX reports."
        ) from _IMPORT_ERROR

    wb = Workbook()
    ws = wb.active
    ws.title = title or "Schedule"

    dates = list(schedule)
    ws.cell(row=1, column=1, value="Employee").font = HEADER_FONT
    for idx, day in enumerate(dates, start=2):
        cell = ws.cell(row=1, column=idx, value=day.isoformat())
        cell.font = HEADER_FONT
        cell.alignment = CENTER

    for row_idx, employee in enumerate(employees, start=2):
        ws.cell(row=row_idx, column=1, value=f"{employee.id} {employee.name}").font = HEADER_FONT
        for col_idx, day in enumerate(dates, start=2):
            cell = ws.cell(row=row_idx, column=col_idx, value=schedule.get_code(employee.id, day))
            cell.alignment = CENTER

    path = Path(path)
    wb.save(path)
    return path
