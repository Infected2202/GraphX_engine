from __future__ import annotations
from datetime import date
from typing import Dict, List, Tuple, Optional, TYPE_CHECKING
import csv
import os
from collections import defaultdict

import pairing

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Внутренняя таблица кодов — устанавливается из app через set_code_map()
_CODE_MAP: Dict[str, str] = {}

if TYPE_CHECKING:
    from production_calendar import ProductionCalendar


def set_code_map(code_map: Dict[str, str]):
    global _CODE_MAP
    _CODE_MAP = dict(code_map)


def _code_of(shift_key: str) -> str:
    return _CODE_MAP.get(shift_key, shift_key)


# Палитра под заданные правила
FILL_NONE = PatternFill(fill_type=None)
FILL_WHITE = PatternFill("solid", fgColor="FFFFFF")
FILL_GRAY = PatternFill("solid", fgColor="DDDDDD")
FILL_M8 = PatternFill("solid", fgColor="00BFFF")  # голубой
FILL_E8 = PatternFill("solid", fgColor="00FF00")  # салатовый
FILL_N8 = PatternFill("solid", fgColor="000000")
FILL_VAC = PatternFill("solid", fgColor="FEC97F")
FILL_WEEKEND = PatternFill("solid", fgColor="E2F0D9")

B_THIN = Border(
    left=Side(style="thin", color="DDDDDD"),
    right=Side(style="thin", color="DDDDDD"),
    top=Side(style="thin", color="DDDDDD"),
    bottom=Side(style="thin", color="DDDDDD"),
)

HEADER_FONT = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")


def _style_for(code: str):
    """Возвращает (Font, Fill) согласно ТЗ."""
    c = (code or "").upper()
    office = None
    if c.endswith("A"):
        office = "A"
    elif c.endswith("B"):
        office = "B"

    # Цвет текста по офисам (A→чёрный, B→красный)
    font_color = "000000"
    if office == "B":
        font_color = "FF0000"

    fill = FILL_NONE

    if c in {"DA", "DB"}:
        fill = FILL_NONE
    elif c in {"NA", "NB", "N4A", "N4B"}:
        fill = FILL_GRAY
    elif c in {"M8A", "M8B"}:
        fill = FILL_M8
    elif c in {"E8A", "E8B"}:
        fill = FILL_E8
    elif c in {"N8A", "N8B"}:
        fill = FILL_N8
        if office == "A":
            font_color = "FFFFFF"
    elif c.startswith("VAC"):
        fill = FILL_VAC
        font_color = "000000"
    elif c == "OFF":
        fill = FILL_NONE
        font_color = "E2F0D9"
    else:
        # Неизвестные коды — зелёный текст для привлечения внимания.
        font_color = "008000"

    return Font(color=font_color), fill


# ------------------- Публичные точки -------------------
def write_workbook(
    path: str,
    ym: str,
    employees: List,
    schedule: Dict[date, List],
    *,
    calendar: "ProductionCalendar" | None = None,
):
    """Совместимая точка: сохраняет XLSX с единственным листом-сеткой."""
    return write_excel_grid(path, ym, employees, schedule, calendar=calendar)


def write_excel_grid(
    path: str,
    ym: str,
    employees: List,
    schedule: Dict[date, List],
    *,
    calendar: "ProductionCalendar" | None = None,
):
    """Единственный лист Excel: сетка сотрудники×дни, ячейки оформлены по правилам."""
    wb = Workbook()
    ws = wb.active
    ws.title = ym
    _write_grid(ws, ym, employees, schedule, calendar=calendar)
    wb.save(path)
    return path


def write_csv_grid(path: str, ym: str, employees: List, schedule: Dict[date, List]):
    """CSV-сетка для простого анализа: первая колонка — сотрудник, первая строка — даты (числа)."""
    dates = sorted(schedule.keys())
    employees_sorted = sorted(employees, key=lambda e: e.id)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Сотрудник"] + [d.day for d in dates])
        for e in employees_sorted:
            row = [f"{e.id} — {e.name}"]
            for d in dates:
                code = ""
                for r in schedule[d]:
                    if r.employee_id == e.id:
                        code = _code_of(r.shift_key)
                        break
                row.append(code)
            w.writerow(row)
    return path


# ------------------- Метрики -------------------
def write_metrics_days_csv(path: str, schedule: Dict[date, List]):
    """По датам: количества DA/DB/M8/NA/NB/N8 (N4/N8 учитываем как ночные; дополнительно считаем N8 отдельно)."""
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["date", "DA", "DB", "M8", "NA", "NB", "N8"])
        for d in sorted(schedule.keys()):
            da = db = m8 = na = nb = n8 = 0
            for a in schedule[d]:
                code = _code_of(a.shift_key).upper()
                if code == "DA":
                    da += 1
                elif code == "DB":
                    db += 1
                elif code in {"M8A", "M8B"}:
                    m8 += 1
                elif code in {"NA", "N4A", "N8A"}:
                    na += 1
                    if code == "N8A":
                        n8 += 1
                elif code in {"NB", "N4B", "N8B"}:
                    nb += 1
                    if code == "N8B":
                        n8 += 1
            w.writerow([d.isoformat(), da, db, m8, na, nb, n8])
    return path


def write_metrics_employees_csv(path: str, employees: List, schedule: Dict[date, List]):
    """По сотрудникам: суммарные часы и количество D/N/O (VAC → O)."""

    def tok(code: str) -> str:
        c = (code or "").upper()
        if c in {"DA", "DB", "M8A", "M8B", "E8A", "E8B"}:
            return "D"
        if c in {"NA", "NB", "N4A", "N4B", "N8A", "N8B"}:
            return "N"
        return "O"

    emp_name = {e.id: e.name for e in employees}
    stats = {e.id: {"hours": 0, "D": 0, "N": 0, "O": 0} for e in employees}
    known_ids = set(stats.keys())

    for d, rows in schedule.items():
        per_emp: Dict[str, Tuple[str, int]] = {}
        for a in rows:
            if a.employee_id not in known_ids:
                continue
            per_emp[a.employee_id] = (_code_of(a.shift_key), a.effective_hours)
        for eid, (code, hrs) in per_emp.items():
            stats[eid]["hours"] += int(hrs)
            stats[eid][tok(code)] += 1

    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["employee_id", "employee", "hours_total", "days_D", "nights_N", "offs_O"])
        for eid in sorted(stats.keys()):
            s = stats[eid]
            w.writerow([eid, emp_name[eid], s["hours"], s["D"], s["N"], s["O"]])
    return path


def _build_norms_summary(employees: List, schedule: Dict[date, List], norm_info: Dict) -> Dict:
    norm_hours = int(norm_info.get("norm_hours") or 0)
    monthly_allowance = int(norm_info.get("monthly_allowance") or 0)
    monthly_cap = int(norm_info.get("monthly_cap") or (norm_hours + monthly_allowance if norm_hours else 0))
    yearly_cap = int(norm_info.get("yearly_cap") or 0)

    hours_by_emp: Dict[str, int] = {}
    for rows in schedule.values():
        for assn in rows:
            hours_by_emp[assn.employee_id] = hours_by_emp.get(assn.employee_id, 0) + int(assn.effective_hours)

    rows_summary: List[Dict[str, object]] = []
    warnings: List[str] = []
    employees_by_id = {e.id: e for e in employees}

    for eid, emp in sorted(employees_by_id.items()):
        total_hours = hours_by_emp.get(eid, 0)
        delta = total_hours - norm_hours if norm_hours else None
        overtime_month = max(0, delta) if (delta is not None) else 0
        yearly_used = emp.ytd_overtime + overtime_month
        yearly_left = yearly_cap - yearly_used if yearly_cap else None
        rows_summary.append(
            {
                "employee_id": eid,
                "employee": emp,
                "hours": total_hours,
                "delta": delta,
                "overtime_month": overtime_month,
                "yearly_left": yearly_left,
            }
        )

        exceeds_month = bool(norm_hours and monthly_cap and total_hours > monthly_cap)
        exceeds_year = bool(yearly_cap and yearly_left is not None and yearly_left < 0)
        if exceeds_month or exceeds_year:
            over_month = total_hours - norm_hours if norm_hours else total_hours
            if exceeds_year and not exceeds_month:
                msg = (
                    f"{eid} — {emp.name}: превышен годовой лимит на {abs(yearly_left)}ч"
                    if yearly_left is not None
                    else f"{eid} — {emp.name}: превышен годовой лимит"
                )
            else:
                leftover = yearly_left if yearly_left is not None else "N/A"
                msg = f"{eid} — {emp.name}: перелимит {over_month}ч; остаток по году {leftover}ч"
            warnings.append(msg)

    return {
        "rows": rows_summary,
        "warnings": warnings,
        "norm_hours": norm_hours,
        "monthly_allowance": monthly_allowance,
        "monthly_cap": monthly_cap,
        "yearly_cap": yearly_cap,
    }


def write_norms_report(
    path: str,
    ym: str,
    employees: List,
    schedule: Dict[date, List],
    norm_info: Dict,
):
    summary = _build_norms_summary(employees, schedule, norm_info)
    os.makedirs(os.path.dirname(path), exist_ok=True)

    norm_hours = summary["norm_hours"]
    monthly_allowance = summary["monthly_allowance"]
    yearly_cap = summary["yearly_cap"]
    warnings = summary["warnings"]
    operations = norm_info.get("operations", []) or []

    with open(path, "w", encoding="utf-8") as f:
        f.write(f"Норма месяца {ym}: {norm_hours}ч\n")
        f.write(f"Допустимое превышение в месяц: +{monthly_allowance}ч\n")
        f.write(f"Допустимый перелимит в год: {yearly_cap}ч\n\n")

        f.write("Фактические часы по сотрудникам:\n")
        for row in summary["rows"]:
            emp = row["employee"]
            hours = row["hours"]
            delta = row["delta"]
            yearly_left = row["yearly_left"]
            if delta is None:
                f.write(f"- {emp.id} — {emp.name}: {hours}ч\n")
            else:
                sign = "+" if delta >= 0 else ""
                f.write(
                    f"- {emp.id} — {emp.name}: {hours}ч (норма {sign}{delta}ч, остаток по году: {yearly_left if yearly_left is not None else 'N/A'}ч)\n"
                )

        f.write("\nСокращения смен:\n")
        if operations:
            for op in sorted(operations, key=lambda x: (x["date"], x["employee_id"])):
                dt = op["date"].isoformat() if hasattr(op.get("date"), "isoformat") else op.get("date")
                f.write(
                    f"- {dt} {op['employee_id']}: {op['from_code']}→{op['to_code']} ({op.get('hours_delta', 0)}ч)\n"
                )
        else:
            f.write("- нет\n")

        f.write("\nПредупреждения:\n")
        if warnings:
            for msg in warnings:
                f.write(f"- {msg}\n")
        else:
            f.write("- нет\n")

    return path, warnings, summary


# ------------------- Пары -------------------
def write_pairs_csv(path: str, pairs: List[Tuple[str,str,int,int]], employees: List):
    """pairs: [(eid1,eid2,overlap_day,overlap_night), ...]"""
    name = {e.id: e.name for e in employees}
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["emp1_id","emp1","emp2_id","emp2","overlap_day","overlap_night"])
        for e1, e2, od, on in pairs:
            w.writerow([e1, name.get(e1,""), e2, name.get(e2,""), od, on])
    return path
# ------------------- Текстовые отчёты по парам -------------------


def write_pairs_text_report(
    out_dir: str,
    ym: str,
    *,
    threshold_day: int,
    window_days: int,
    max_ops: int,
    hours_budget: int,
    prev_pairs: Optional[List[Tuple[str, str, int, int]]],
    curr_pairs: List[Tuple[str, str, int, int]],
    prev_days_total: Optional[int],
    curr_days_total: Optional[int],
    ops_log: List[str],
    apply_log: Optional[List[str]] = None,
    pair_score_before: int,
    pair_score_after: int,
) -> str:
    """Пишет текстовый отчёт по парам и возвращает путь к файлу."""

    os.makedirs(out_dir, exist_ok=True)
    path = os.path.join(out_dir, f"{ym}_pairs.txt")

    def fmt_pairs(lst: List[Tuple[str, str, int, int]], top: int = 10) -> List[str]:
        rows: List[str] = []
        for e1, e2, d, n in lst[:top]:
            rows.append(f" {e1}~{e2} d{d}/n{n}")
        return rows

    prev_excl = pairing.exclusive_matching_by_day(prev_pairs or [], threshold_day=threshold_day)
    curr_excl = pairing.exclusive_matching_by_day(curr_pairs, threshold_day=threshold_day)

    def key_of(a: str, b: str) -> str:
        return f"{a}~{b}" if a < b else f"{b}~{a}"

    prev_set = {key_of(a, b) for a, b, _, _ in prev_excl}
    curr_set = {key_of(a, b) for a, b, _, _ in curr_excl}
    retained = sorted(prev_set & curr_set)
    broken = sorted(prev_set - curr_set)
    new = sorted(curr_set - prev_set)

    def bucket_counts(lst: List[Tuple[str, str, int, int]]) -> Dict[str, int]:
        buckets = {"0..3": 0, "4..7": 0, f"{threshold_day}..": 0}
        for _, _, d, _ in lst:
            if d <= 3:
                buckets["0..3"] += 1
            elif d <= 7:
                buckets["4..7"] += 1
            else:
                buckets[f"{threshold_day}.."] += 1
        return buckets

    curr_buckets = bucket_counts(curr_excl)

    def by_emp(lst: List[Tuple[str, str, int, int]]) -> Dict[str, Tuple[str, int]]:
        data: Dict[str, Tuple[str, int]] = {}
        for a, b, d, _ in lst:
            data[a] = (b, d)
            data[b] = (a, d)
        return data

    prev_emp = by_emp(prev_excl)
    curr_emp = by_emp(curr_excl)

    with open(path, "w", encoding="utf-8") as f:
        pct: Optional[int] = None
        if curr_days_total and curr_days_total > 0:
            pct = int(round(100.0 * threshold_day / curr_days_total))

        f.write("[pairs.config]\n")
        f.write(f"threshold_day={threshold_day}d")
        if pct is not None:
            f.write(f" (≈{pct}%)")
        f.write(
            f", window_days={window_days}, max_ops={max_ops}, hours_budget={hours_budget}\n\n"
        )

        if prev_pairs is not None:
            f.write("[pairs.prev_month]\n")
            f.write(f"pairs_strong={len(prev_excl)}\n")
            for row in fmt_pairs(prev_excl):
                f.write(row + "\n")
            f.write("\n")

        f.write("[pairs.current_month]\n")
        f.write(f"pairs_strong={len(curr_excl)}\n")
        for row in fmt_pairs(curr_excl):
            f.write(row + "\n")
        f.write(
            "day_overlap buckets: "
            f"[0..3]={curr_buckets['0..3']}, [4..7]={curr_buckets['4..7']}, "
            f"[{threshold_day}..]={curr_buckets[f'{threshold_day}..']}\n\n"
        )

        f.write("[pairs.delta]\n")
        f.write(f"retained={len(retained)}, broken={len(broken)}, new={len(new)}\n")
        if retained:
            f.write(" retained: " + ", ".join(retained) + "\n")
        if broken:
            f.write(" broken  : " + ", ".join(broken) + "\n")
        if new:
            f.write(" new     : " + ", ".join(new) + "\n")
        f.write(
            f"pair_score: {pair_score_before} → {pair_score_after} (Δ={pair_score_after - pair_score_before})\n\n"
        )

        f.write("[pairs.by_employee]\n")
        emp_ids = sorted(set(prev_emp.keys()) | set(curr_emp.keys()))
        for eid in emp_ids:
            prev_info = prev_emp.get(eid)
            curr_info = curr_emp.get(eid)
            prev_str = f"{prev_info[0]}(d{prev_info[1]})" if prev_info else "-"
            curr_str = f"{curr_info[0]}(d{curr_info[1]})" if curr_info else "-"
            f.write(f"{eid}: prev={prev_str}, curr={curr_str}\n")
        f.write("\n")

        f.write("[pair_breaking.summary]\n")
        accepted = sum(1 for line in (apply_log or []) if "-> ACCEPT" in line)
        f.write(f"ops_applied≈{accepted}\n\n")

        f.write("[pair_breaking.ops]\n")
        filtered_ops = [
            ln for ln in (ops_log or []) if "-> ACCEPT" not in ln and "-> REJECT" not in ln
        ]
        for line in filtered_ops:
            f.write(line + "\n")

    return path


def render_pairs_text_block(
    *,
    ym: str,
    apply_log: List[str],
    threshold_day: int,
    window_days: int,
    max_ops: int,
    hours_budget: int,
    prev_pairs: Optional[List[Tuple[str, str, int, int]]],
    curr_pairs: List[Tuple[str, str, int, int]],
    curr_days_total: Optional[int],
    ops_log: List[str],
    pair_score_before: int,
    pair_score_after: int,
) -> str:
    """Возвращает текстовый блок отчёта по парам (без записи на диск)."""

    def fmt_pairs(lst: List[Tuple[str, str, int, int]], top: int = 10) -> List[str]:
        return [f" {a}~{b} d{d}/n{n}" for a, b, d, n in lst[:top]]

    prev_excl = pairing.exclusive_matching_by_day(prev_pairs or [], threshold_day=threshold_day)
    curr_excl = pairing.exclusive_matching_by_day(curr_pairs, threshold_day=threshold_day)

    def key_of(a: str, b: str) -> str:
        return f"{a}~{b}" if a < b else f"{b}~{a}"

    prev_set = {key_of(a, b) for a, b, _, _ in prev_excl}
    curr_set = {key_of(a, b) for a, b, _, _ in curr_excl}
    retained = sorted(prev_set & curr_set)
    broken = sorted(prev_set - curr_set)
    new = sorted(curr_set - prev_set)

    def bucket_counts(lst: List[Tuple[str, str, int, int]]) -> Dict[str, int]:
        buckets = {"0..3": 0, "4..7": 0, f"{threshold_day}..": 0}
        for _, _, d, _ in lst:
            if d <= 3:
                buckets["0..3"] += 1
            elif d <= 7:
                buckets["4..7"] += 1
            else:
                buckets[f"{threshold_day}.."] += 1
        return buckets

    buckets = bucket_counts(curr_excl)

    pct = None
    if curr_days_total and curr_days_total > 0:
        pct = int(round(100.0 * threshold_day / curr_days_total))

    lines: List[str] = []
    lines.append("[pairs.config]")
    cfg_line = f"threshold_day={threshold_day}d"
    if pct is not None:
        cfg_line += f" (≈{pct}%)"
    cfg_line += f", window_days={window_days}, max_ops={max_ops}, hours_budget={hours_budget}"
    lines.append(cfg_line)
    lines.append("")

    if prev_pairs is not None:
        lines.append("[pairs.prev_month]")
        lines.append(f"pairs_strong={len(prev_excl)}")
        lines += fmt_pairs(prev_excl)
        lines.append("")

    lines.append("[pairs.current_month]")
    lines.append(f"pairs_strong={len(curr_excl)}")
    lines += fmt_pairs(curr_excl)
    lines.append(
        f"day_overlap buckets: [0..3]={buckets['0..3']}, [4..7]={buckets['4..7']}, [{threshold_day}..]={buckets[f'{threshold_day}..']}"
    )
    lines.append("")

    lines.append("[pairs.delta]")
    lines.append(f"retained={len(retained)}, broken={len(broken)}, new={len(new)}")
    if retained:
        lines.append(" retained: " + ", ".join(retained))
    if broken:
        lines.append(" broken  : " + ", ".join(broken))
    if new:
        lines.append(" new     : " + ", ".join(new))
    lines.append(f"pair_score: {pair_score_before} → {pair_score_after} (Δ={pair_score_after - pair_score_before})")
    lines.append("")

    if apply_log:
        lines.append("[pair_breaking.apply]")
        lines += apply_log
        lines.append("")

    lines.append("[pair_breaking.summary]")
    accepted = sum(1 for line in (apply_log or []) if "-> ACCEPT" in line)
    lines.append(f"ops_applied≈{accepted}")
    lines.append("")

    lines.append("[pair_breaking.ops]")
    filtered_ops = [ln for ln in (ops_log or []) if "-> ACCEPT" not in ln and "-> REJECT" not in ln]
    lines += filtered_ops
    lines.append("")

    return "\n".join(lines)


def append_pairs_to_log(
    out_dir: str,
    ym: str,
    apply_log: List[str],
    **kwargs,
) -> str:
    """Апендит блок pairs-отчёта в <out_dir>/<ym>_log.txt и возвращает путь."""

    os.makedirs(out_dir, exist_ok=True)
    log_path = os.path.join(out_dir, f"{ym}_log.txt")
    block = render_pairs_text_block(ym=ym, apply_log=apply_log, **kwargs)
    with open(log_path, "a", encoding="utf-8") as f:
        f.write("\n" + block)
    return log_path

# ------------------- Логи -------------------
def write_log_txt(path: str, lines: List[str]):
    with open(path, "w", encoding="utf-8") as f:
        for ln in lines:
            f.write(ln.rstrip() + "\n")
    return path


# ------------------- ГРИД -------------------

def _is_weekend_or_off(dt: date, calendar: "ProductionCalendar" | None) -> bool:
    if calendar and calendar.is_working_override(dt):
        return False
    if dt.weekday() >= 5:
        return True
    if calendar and calendar.is_off_date(dt):
        return True
    return False


def _write_grid(
    ws,
    ym: str,
    employees: List,
    schedule: Dict[date, List],
    *,
    calendar: "ProductionCalendar" | None = None,
):
    # Заголовок: первая строка — номера дней; A1 — «Сотрудник»
    dates = sorted(schedule.keys())
    weekend_flags = {d: _is_weekend_or_off(d, calendar) for d in dates}
    ws.cell(row=1, column=1, value="Сотрудник")
    for j, d in enumerate(dates, start=2):
        cell = ws.cell(row=1, column=j, value=d.day)
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = B_THIN
        if weekend_flags.get(d):
            cell.fill = FILL_WEEKEND

    # Тело
    employees_sorted = sorted(employees, key=lambda e: e.id)
    for i, e in enumerate(employees_sorted, start=2):
        c = ws.cell(row=i, column=1, value=f"{e.id} — {e.name}")
        c.font = HEADER_FONT
        c.alignment = LEFT
        c.border = B_THIN
        for j, d in enumerate(dates, start=2):
            code = ""
            for r in schedule[d]:
                if r.employee_id == e.id:
                    code = _code_of(r.shift_key)
                    break
            cell = ws.cell(row=i, column=j, value=code)
            cell.alignment = CENTER
            cell.border = B_THIN
            font, fill = _style_for(code)
            code_upper = (code or "").upper()
            if weekend_flags.get(d):
                if fill is FILL_NONE:
                    fill = FILL_WEEKEND
                elif code_upper in {"", "OFF"}:
                    fill = FILL_WEEKEND
            cell.font = font
            cell.fill = fill

    # Вёрстка
    ws.freeze_panes = "B2"
    ws.column_dimensions['A'].width = 28
    for col in range(2, 2 + len(dates)):
        ws.column_dimensions[get_column_letter(col)].width = 6
    # Фильтр не включаем, чтобы не вносить лишних артефактов формата
